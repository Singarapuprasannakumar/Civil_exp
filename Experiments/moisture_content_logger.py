import os
from openpyxl import Workbook, load_workbook

EXCEL_FILE = "moisture_content_results.xlsx"


def create_excel_file():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Moisture Content"

        ws.append([
            "Observation No",
            "Container No",
            "Mass of Container + Lid (M1) (g)",
            "Mass of Container + Lid + Moist Soil (M2) (g)",
            "Mass of Container + Lid + Oven Dry Soil (M3) (g)",
            "Mass of Water Mw (g)",
            "Mass of Oven Dry Soil Md (g)",
            "Moisture Content (%)"
        ])

        wb.save(EXCEL_FILE)


def append_to_excel(data):
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Moisture Content"]

    for row in data:
        ws.append(row)

    wb.save(EXCEL_FILE)


def moisture_content_test():
    print("\n" + "=" * 60)
    print("MOISTURE CONTENT TEST")
    print("=" * 60)

    n = int(input("\nEnter number of observations: "))

    results = []

    print("\nEnter observations:\n")

    for i in range(1, n + 1):
        print(f"\nObservation {i}")

        container_no = input("Container Number: ")

        M1 = float(input("Mass of container + lid, M1 (g): "))
        M2 = float(input("Mass of container + lid + moist soil, M2 (g): "))
        M3 = float(input("Mass of container + lid + oven dried soil, M3 (g): "))

        Mw = M2 - M3
        Md = M3 - M1

        if Md == 0:
            moisture = 0
        else:
            moisture = (Mw / Md) * 100

        print("\nCalculated Results")
        print("------------------------")
        print(f"Mass of water (Mw)           = {Mw:.2f} g")
        print(f"Mass of oven dried soil (Md) = {Md:.2f} g")
        print(f"Moisture content             = {moisture:.2f} %")

        results.append([
            i,
            container_no,
            M1,
            M2,
            M3,
            round(Mw, 2),
            round(Md, 2),
            round(moisture, 2)
        ])

    append_to_excel(results)

    avg_moisture = sum(row[7] for row in results) / len(results)

    print("\n" + "=" * 60)
    print("FINAL RESULTS")
    print("=" * 60)

    print("\nObservation Summary")

    for row in results:
        print(
            f"Obs-{row[0]} | "
            f"Container: {row[1]} | "
            f"Moisture Content = {row[7]:.2f} %"
        )

    print(f"\nAverage Moisture Content = {avg_moisture:.2f} %")

    print(f"\nResults appended to Excel file: {EXCEL_FILE}")


def main():
    create_excel_file()

    while True:
        moisture_content_test()

        choice = int(
            input(
                "\nDo you want to continue?\n"
                "Enter 1 for Yes and 0 for No: "
            )
        )

        if choice == 0:
            print("\nProgram terminated successfully.")
            break

        elif choice != 1:
            print("\nInvalid choice. Program terminated.")
            break


if __name__ == "__main__":
    main()
